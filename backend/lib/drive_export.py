"""
drive_export.py — XLSX build + Drive upload. Port of driveExport.ts.
"""
from __future__ import annotations
import io
from typing import Optional
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from .config import DRIVE_EXPORT_FOLDER, EXPORT_LABELS
from .googleauth import drive_client


def build_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Placements"

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    last_row = ws.max_row
    for label in [EXPORT_LABELS["relevance"], EXPORT_LABELS["seo"]]:
        if label in headers:
            col_idx = headers.index(label) + 1
            col_letter = get_column_letter(col_idx)
            rule = ColorScaleRule(
                start_type="min", start_color="8B0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                end_type="max", end_color="006400",
            )
            if last_row > 1:
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{last_row}", rule)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _get_or_create_folder(name: str) -> str:
    drive = await drive_client()
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    resp = drive.files().list(q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute()
    files = resp.get("files", [])
    if files:
        return files[0]["id"]
    created = drive.files().create(
        body={"name": name, "mimeType": "application/vnd.google-apps.folder"},
        fields="id", supportsAllDrives=True,
    ).execute()
    return created["id"]


async def export_to_drive(rows: list[dict], filename: str, fmt: str = "excel") -> str:
    import re as _re
    is_csv = fmt == "csv" or filename.lower().endswith(".csv")
    name = filename if _re.search(r"\.(xlsx|csv)$", filename, _re.IGNORECASE) else f"{filename}.xlsx"

    if is_csv:
        hdrs = list(rows[0].keys()) if rows else []
        def esc(v): return f'"{str(v).replace(chr(34), chr(34)*2)}"'
        lines = [",".join(hdrs)] + [",".join(esc(r[h]) for h in hdrs) for r in rows]
        body = "\n".join(lines).encode("utf-8")
        mime = "text/csv"
    else:
        body = build_xlsx(rows)
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    folder_id = await _get_or_create_folder(DRIVE_EXPORT_FOLDER)
    drive = await drive_client()

    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(body), mimetype=mime, resumable=False)
    uploaded = drive.files().create(
        body={"name": name, "parents": [folder_id]},
        media_body=media,
        fields="id,webViewLink",
        supportsAllDrives=True,
    ).execute()
    return uploaded.get("webViewLink", "")
